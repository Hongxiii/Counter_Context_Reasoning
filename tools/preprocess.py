import pandas as pd
import json
import argparse
import openpyxl
import sys


def process(annotation_file, dataset_file, args):
    # 读取标注
    input_data = pd.ExcelFile(annotation_file)
    if args.task == "caption":
        sheet_data = input_data.parse(input_data.sheet_names[0])
        sheet_data = sheet_data.iloc[:, [0, 1, 3, 5, 7, 9]]
        sheet_data = sheet_data.apply(lambda row: list(row.dropna().astype(str)), axis=1).tolist()
    elif args.task == "explanation":
        sheet_data = input_data.parse(input_data.sheet_names[1])
        sheet_data = sheet_data.iloc[:, [0, 1, 3, 5, 7, 9]]
        sheet_data = sheet_data.apply(lambda row: list(row.dropna().astype(str)), axis=1).tolist()
    else:
        sheet_data = input_data.parse(input_data.sheet_names[2])
        sheet_data = sheet_data.iloc[:, [0, 1, 2, 4, 5, 7, 8, 10, 11, 13, 14]]
        sheet_data = sheet_data.apply(lambda row: list(row.dropna().astype(str)), axis=1).tolist()
    # 根据id拆分正负样本
    true_data = []
    false_data = []
    for data in sheet_data:
        if data[0][2] == "t":
            true_data.append(data)
        else:
            false_data.append(data)
    # 转存标注
    if args.task == "caption":
        column = 6
    elif args.task == "explanation":
        column = 7
    else:
        column = 8
    sheet = openpyxl.load_workbook(dataset_file)
    true_sheet, false_sheet = sheet.worksheets[1], sheet.worksheets[2]
    for data in true_data:
        data_id = data[0]
        for cell in true_sheet['A']:
            if cell.value == data_id:
                crowd_caption = ""
                for d in data[1:]:
                    crowd_caption += '[' + d + ']'
                crowd_caption_cell = true_sheet.cell(row=cell.row, column=column)
                crowd_caption_cell.value = crowd_caption
                break
    for data in false_data:
        data_id = data[0]
        for cell in false_sheet['A']:
            if cell.value == data_id:
                crowd_caption = ""
                for d in data[1:]:
                    crowd_caption += '[' + d + ']'
                crowd_caption_cell = false_sheet.cell(row=cell.row, column=column)
                crowd_caption_cell.value = crowd_caption
                break
    sheet.save(dataset_file)
    

if __name__ == '__main__':
    # 获取控制台参数
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-task", type=str, required=True, choices=["caption","vqa","explanation"]
    )
    args = parser.parse_args()
    # 读取文件
    annotation_file = '../dataset/annotation.xlsx'
    dataset_file = '../dataset/dataset.xlsx'
    # 处理数据
    process(annotation_file, dataset_file, args)